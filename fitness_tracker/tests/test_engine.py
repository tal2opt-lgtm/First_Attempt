"""
Tests for the normaliser and the workbook appender.

Run:  cd fitness_tracker/engine && python -m pytest ../tests -q
(or)  cd fitness_tracker/tests && python test_engine.py
"""

import shutil
import sys
from pathlib import Path

ENGINE = Path(__file__).resolve().parents[1] / "engine"
TEMPLATE = Path(__file__).resolve().parents[1] / "template" / "fitness_master.xlsx"
sys.path.insert(0, str(ENGINE))

import openpyxl  # noqa: E402
from schema import normalise, parse_date, parse_pace, parse_duration, parse_time_range  # noqa: E402
from append_row import append_record, is_duplicate  # noqa: E402


# What the vision model is expected to return for the two provided screenshots
# (IMG_0637 summary + IMG_0638 map). Elevation is null: it is not on either view.
FOUR_AUG_RAW = {
    "date": "Tue, 4 Aug",
    "time_range": "5:17-6:17",
    "duration": "1:00:04",
    "distance_km": 8.65,
    "avg_heart_rate": 160,
    "avg_pace": "6:56",
    "elevation_gain": None,
    "avg_power": 151,
    "avg_cadence": 161,
    "temperature": 25,
}


def test_small_parsers():
    assert parse_date("Tue, 4 Aug", 2026) == "4.8.26"
    assert parse_date("9 Jun 2026", 2000) == "9.6.26"
    assert parse_time_range("5:17–6:17") == "05:17-06:17"
    assert parse_duration("1:00:04") == "1:00:04"
    assert parse_duration("30:07") == "0:30:07"
    assert parse_pace("6'56''/KM") == "6:56"


def test_normalise_four_aug():
    rec = normalise(FOUR_AUG_RAW, default_year=2026)
    assert rec.as_row() == [
        "4.8.26", "05:17-06:17", "1:00:04", 8.65, 160, "6:56", "X", 151, 161, 25,
    ]


def test_append_matches_existing_columns(tmp_path=None):
    import tempfile
    tmp = Path(tempfile.mkdtemp())
    work = tmp / "master.xlsx"
    shutil.copy(TEMPLATE, work)

    rec = normalise(FOUR_AUG_RAW, default_year=2026)
    row = append_record(work, rec)

    wb = openpyxl.load_workbook(work)
    ws = wb["ריצות"]
    written = [ws.cell(row=row, column=c).value for c in range(1, 11)]
    assert written == rec.as_row(), written

    # The nutrition sheet and its formulas must survive untouched.
    assert "תזונה" in wb.sheetnames
    nut = wb["תזונה"]
    assert str(nut["I2"].value).startswith("=")  # XLOOKUP formula intact

    # Duplicate guard.
    assert is_duplicate(work, rec) is True
    print(f"OK: appended to row {row}: {written}")


if __name__ == "__main__":
    test_small_parsers()
    test_normalise_four_aug()
    test_append_matches_existing_columns()
    print("All tests passed.")
