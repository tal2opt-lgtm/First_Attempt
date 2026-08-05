#!/usr/bin/env python3
"""
xlsx_to_csv.py — export the existing runs from the master workbook into the
canonical CSV format that the iPhone Shortcut appends to. Run once so the living
CSV log starts with all of your history already in it.

Usage:
    python xlsx_to_csv.py --master ../template/fitness_master.xlsx --out ../Runs.csv
"""

from __future__ import annotations

import argparse
import csv
import datetime as dt
from pathlib import Path

import openpyxl

from schema import TARGET_SHEET_NAME, MISSING


def _duration(v) -> str:
    if isinstance(v, dt.time):
        return f"{v.hour}:{v.minute:02d}:{v.second:02d}"
    if isinstance(v, dt.timedelta):
        total = int(v.total_seconds())
        h, rem = divmod(total, 3600)
        m, s = divmod(rem, 60)
        return f"{h}:{m:02d}:{s:02d}"
    return str(v).strip()


def _pace(v) -> str:
    # Stored as a clock time where hour=minutes and minute=seconds of the pace.
    if isinstance(v, dt.time):
        return f"{v.hour}:{v.minute:02d}"
    return str(v).strip()


def _plain(v) -> str:
    if v is None:
        return MISSING
    if isinstance(v, float) and v.is_integer():
        return str(int(v))
    return str(v).strip()


def export(master: Path, out: Path) -> int:
    wb = openpyxl.load_workbook(master)
    ws = wb[TARGET_SHEET_NAME]
    n = 0
    with out.open("w", encoding="utf-8-sig", newline="") as fh:
        w = csv.writer(fh)
        for r in range(2, ws.max_row + 1):
            row = [ws.cell(row=r, column=c).value for c in range(1, 11)]
            if row[0] in (None, ""):
                continue
            w.writerow([
                _plain(row[0]),        # date
                _plain(row[1]),        # time
                _duration(row[2]),     # duration
                _plain(row[3]),        # distance
                _plain(row[4]),        # hr
                _pace(row[5]),         # pace
                _plain(row[6]),        # elevation (X kept)
                _plain(row[7]),        # power
                _plain(row[8]),        # cadence
                _plain(row[9]),        # temperature
            ])
            n += 1
    return n


def main(argv=None) -> int:
    ap = argparse.ArgumentParser(description="Export master runs to canonical CSV.")
    ap.add_argument("--master", required=True)
    ap.add_argument("--out", required=True)
    args = ap.parse_args(argv)
    n = export(Path(args.master), Path(args.out))
    print(f"Exported {n} run(s) to {args.out}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
