"""
append_row.py
=============
Append one RunRecord as a new row to the master workbook's "ריצות" sheet,
directly under the last existing row, while preserving:

  * the second sheet ("תזונה") and all of its XLOOKUP formulas,
  * column widths / number formats / styles of the runs sheet,
  * the exact header order.

We copy the style of the previous data row onto the new one so the appended
row looks identical to the ones the user typed by hand.
"""

from __future__ import annotations

import copy
from pathlib import Path

import openpyxl

from schema import HEADERS, TARGET_SHEET_NAME, RunRecord, ExtractionError


def _validate_headers(ws) -> None:
    actual = [ws.cell(row=1, column=i + 1).value for i in range(len(HEADERS))]
    if actual != HEADERS:
        raise ExtractionError(
            "Header mismatch in target sheet.\n"
            f"  expected: {HEADERS}\n"
            f"  found:    {actual}"
        )


def _last_data_row(ws) -> int:
    """Row index of the last row that has anything in column A."""
    last = 1  # header is row 1
    for r in range(2, ws.max_row + 1):
        if ws.cell(row=r, column=1).value not in (None, ""):
            last = r
    return last


def append_record(master_path: str | Path, record: RunRecord,
                  *, dry_run: bool = False) -> int:
    """
    Append `record` to the workbook at `master_path`.
    Returns the row number that was written. With dry_run=True nothing is saved.
    """
    master_path = Path(master_path)
    wb = openpyxl.load_workbook(master_path)
    if TARGET_SHEET_NAME not in wb.sheetnames:
        raise ExtractionError(
            f"sheet {TARGET_SHEET_NAME!r} not found (have: {wb.sheetnames})"
        )
    ws = wb[TARGET_SHEET_NAME]
    _validate_headers(ws)

    last = _last_data_row(ws)
    new_row = last + 1
    values = record.as_row()

    for col_idx, value in enumerate(values, start=1):
        cell = ws.cell(row=new_row, column=col_idx, value=value)
        # Inherit the look of the row above so the table stays visually uniform.
        if last >= 2:
            src = ws.cell(row=last, column=col_idx)
            if src.has_style:
                cell._style = copy.copy(src._style)

    if not dry_run:
        wb.save(master_path)
    return new_row


def is_duplicate(master_path: str | Path, record: RunRecord) -> bool:
    """
    True if a row with the same Date AND Time already exists — so re-sharing the
    same screenshot twice does not create a duplicate line.
    """
    wb = openpyxl.load_workbook(master_path, read_only=True)
    ws = wb[TARGET_SHEET_NAME]
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        d, t = row[0], row[1]
        if d is None:
            continue
        if str(d).strip() == record.date and str(t).strip() == record.time_range:
            return True
    return False
